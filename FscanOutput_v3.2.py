#coding=utf-8
import re
import time
import traceback
import openpyxl
import sys
import os
import chardet
import openpyxl as p
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font
from chardet.universaldetector import UniversalDetector
from openpyxl.styles import PatternFill
from urllib.parse import urlparse



def get_encoding(file):
    # 二进制方式读取，获取字节数据，检测类型
    with open(file, 'rb') as f:
        data = f.read()
        return chardet.detect(data)['encoding']

def get_encode_info(file):
    with open(file, 'rb') as f:
        data = f.read()
        result = chardet.detect(data)
        return result['encoding']

def read_file(file):
    with open(file, 'rb') as f:
        return f.read()

def write_file(content, file):
    with open(file, 'wb') as f:
        f.write(content)


def convert_encode2utf8(file, original_encode, des_encode):
    file_content = read_file(file)
    file_decode = file_content.decode(original_encode, 'ignore')
    file_encode = file_decode.encode(des_encode)
    write_file(file_encode, file)


def OpenFile():
    file_name = getInput()
    datalist = []
    datastr = ''
    with open(file_name, encoding='utf-8', errors='ignore') as f:
        for i in f.readlines():
            datalist.append(i.strip())
    with open(file_name, encoding='utf-8', errors='ignore') as f:
        datastr = f.read()

    return datalist, datastr

#输出存活端口
def open_port_export(datalist):
    sheetList = [['ip', 'port', 'protocol']]
    unique_dic = {}

    for i in datalist:
        p = re.findall(r'^\d[^\s]+ open.*?$', i)

        if len(p) != 0:
            p1 = list(p)
            for u in p1:
                ip = re.findall(r"\d+\.\d+\.\d+\.\d+", u)
                port = re.findall("(?<=:)\d+" , u)
                protocol = re.findall(r"open (\S+)$" , u)
                if len(protocol) > 0:
                    protocol = protocol[0]
                else:
                    protocol = ''

                try:
                    ip_c = ".".join(ip[0].split('.')[:3])
                    if unique_dic.get(ip_c):
                        if unique_dic[ip_c].get(ip[0]):
                            unique_dic[ip_c].get(ip[0])[f'{ip[0]}_{port[0]}'] = [ ip[0] , port[0] , protocol]
                        else:
                            unique_dic[ip_c][ip[0]] = {
                                f'{ip[0]}_{port[0]}' : [ ip[0] , port[0] , protocol]
                            }
                    else:
                        unique_dic[ip_c] = {
                            ip[0] : {
                                f'{ip[0]}_{port[0]}' : [ ip[0] , port[0] , protocol]
                            }
                            
                        }
                except Exception as e:
                    traceback.print_exc()
                
    for ip_c in unique_dic:
        ip_c_dic = unique_dic[ip_c]
        
        for ip in ip_c_dic:
            ip_port_dic = ip_c_dic[ip]
            for ip_port in ip_port_dic:
                sheetList.append(ip_port_dic[ip_port]) 


    OutPut('OpenPort', sheetList)

#输出IP段内存货数量
def alive_ip_export(datalist):
    unique_dic = {}
    sheetList = [['IP段', '段存活数量']]

    for t in datalist:
        Ip_d = re.findall(r"\[\*]\sLiveTop\s\d+\.\d+\.\d+\.\d+/\d+.*", t)
        if len(Ip_d) != 0:
            p1 = list(Ip_d)

            for u in p1:
                ip_duan = re.findall(r"\d+\.\d+\.\d+\.\d+/\d+", u)
                No = re.findall(r"\d+$", u)
                try:
                    unique_dic[f'{ip_duan[0]}'] = [ ip_duan[0], No[0] ]
                except:
                    pass
                
    for key in unique_dic:
        sheetList.append(unique_dic[key])
        
    OutPut('AliveIp', sheetList)



def os_scan_export(datalist):
    unique_dic = {}
    replaceList = ["[*]", '\t', "\x01", '\x02', "(", ")"]
    sheetList = [['ip', 'os', 'hostname', 'vuln', 'netcard']]

    for index, t in enumerate(datalist):
        p = re.findall(r"OsInfo(.*)", t)
        find_17010 = re.findall(r"MS17-010(.*)", t)
        find_netbios = re.findall(r"NetBios(.*)", t)
        find_netcard = re.findall(r"NetInfo(.*)", t)
        if len(p) != 0:
            p1 = list(p)
            for u in p1:
                ip = re.findall(r"\d+\.\d+\.\d+\.\d+", u)
                #删除无用字符
                for q in replaceList:
                    u = u.replace(q, "")

                os_version = u.replace(ip[0], '').replace('\t', '').replace('(',"").replace(')',"").strip()
                try:
                    ip_c = ".".join(ip[0].split('.')[:3])
                    
                    if unique_dic.get(ip_c):
                        if unique_dic[ip_c].get(ip[0]):
                            unique_dic[ip_c].get(ip[0])[1] = os_version
                        else:
                            unique_dic[ip_c][ip[0]] = [ ip[0], os_version, None, None, None ]
                    else:
                        unique_dic[ip_c] = {
                            ip[0] : [ ip[0], os_version, None, None, None ]
                        }
                except:
                    pass
                
        if len(find_17010) != 0:
            ms17010_list = list(find_17010)
            for vuln in ms17010_list:
                try:
                    ip = vuln.strip().split()[0]
                    ip_c = ".".join(ip.split('.')[:3])
                    
                    os_version = " ".join(vuln.strip().split(ip)).strip()
                    os_version = os_version.replace('(',"").replace(')',"").replace('\t','').strip()
                    
                    if unique_dic.get(ip_c):
                        if unique_dic[ip_c].get(ip):
                            unique_dic[ip_c].get(ip)[3] = 'MS17-010'
                            if unique_dic[ip_c].get(ip)[1] == None:
                                unique_dic[ip_c].get(ip)[1] = os_version
                            
                        else:
                            unique_dic[ip_c][ip]=  [ip, os_version, None, 'MS17-010', None ]
                    else:
                        unique_dic[ip_c] = {
                            ip : [ip, os_version, None, 'MS17-010', None ]
                        }
                except:
                    traceback.print_exc()
                    
        if len(find_netbios) != 0:
            netbios_list = list(find_netbios)
            for netbios_info_str in netbios_list:
                try:
                    slice = netbios_info_str.strip().split()
                    if slice[1] == '[+]':
                        ip = slice[0]
                        hostname = slice[2]
                        os_version = None
                        if len(slice) > 3:
                            os_version = " ".join(slice[3:])
                            
                        ip_c = ".".join(ip.split('.')[:3])
                        
                        if unique_dic.get(ip_c):
                            if unique_dic[ip_c].get(ip):
                                if unique_dic[ip_c].get(ip)[2] == None:
                                    unique_dic[ip_c].get(ip)[2] = hostname
                                if unique_dic[ip_c].get(ip)[1] == None:
                                    unique_dic[ip_c].get(ip)[1] = os_version
                            else:
                                unique_dic[ip_c][ip] = [ip, os_version, hostname, None, None]
                        else:
                            unique_dic[ip_c] = {
                                ip : [ip, os_version, hostname, None, None]
                            }
                    else:
                        ip = slice[0]
                        hostname = slice[1]
                        os_version = None
                        if len(slice) > 2:
                            os_version = " ".join(slice[2:])
                        
                        if unique_dic.get(ip_c):
                            if unique_dic[ip_c].get(ip):
                                if unique_dic[ip_c].get(ip)[2] == None:
                                    unique_dic[ip_c].get(ip)[2] = hostname
                                if unique_dic[ip_c].get(ip)[1] == None:
                                    unique_dic[ip_c].get(ip)[1] = os_version
                            else:
                                unique_dic[ip_c][ip] = [ip, os_version, hostname, None, None]
                        else:
                            unique_dic[ip_c] = {
                                ip : [ip, os_version, hostname, None, None]
                            }
                except:
                    traceback.print_exc()

        if len(find_netcard) != 0:
            next_line = index + 1
            ip = datalist[next_line].strip().split('[*]')[1].strip()
            card_interface = []
            card_info = None
            while True:
                next_line += 1
                if '[->]' in datalist[next_line]:
                    interface = datalist[next_line].split('[->]')[1].strip()
                    card_interface.append(interface)
                else:
                    break
            if card_interface:
                card_info = ", ".join(card_interface)
            try:
                ip_c = ".".join(ip.split('.')[:3])
                if unique_dic.get(ip_c):
                    if unique_dic[ip_c].get(ip):
                        unique_dic[ip_c].get(ip)[4] = card_info
                    else:
                        unique_dic[ip_c][ip] = [ip, None, None, None, card_info]
                else:
                    unique_dic[ip_c] = {
                        ip : [ip, None, None, None, card_info]
                    }
            except:
                traceback.print_exc()

                    
    for ip_c in unique_dic:
        for ip in unique_dic[ip_c]:
            sheetList.append(unique_dic[ip_c][ip])

    OutPut('OsList', sheetList)


#输出poc漏洞列表
def vulnscan_export(datalist):
    unique_dic = {}
    sheetList = [['target', 'vuln']]

    for i in datalist:
        find_vuln = re.findall(r"PocScan.*", i)

        if len(find_vuln) != 0:
            vuln_str_list = list(find_vuln)
            for vuln_str in vuln_str_list:
                slice = vuln_str.split('PocScan ')[1].split()
                try:
                    target = slice[0]
                    vuln_info = " ".join(slice[1:])
                    unique_dic[f"{target}_{vuln_info}"] = [ target, vuln_info ]
                except:
                    pass
                
    for key in unique_dic:
        sheetList.append(unique_dic[key])
    OutPut('vulnscan', sheetList)


#输出exp漏洞列表
def exp_list_export(datalist):
    unique_dic = {}
    sheetList = [['ip', 'bug_exp']]

    for i in datalist:
        p = re.findall(r"\[\+]\s\d+\.\d+\.\d+\.\d+.*", i)

        # print(p)

        if len(p) != 0:
            p1 = list(p)
            for u in p1:
                ip = re.findall(r"\d+\.\d+\.\d+\.\d+", u)
                bug = u.replace(ip[0], '').replace("[+]", "").replace('\t', '').strip()
                print(bug)
                # ip.append(bug)
                # sheetList.append(ip)
                try:
                    unique_dic[f'{ip[0]}_{bug}'] = [ ip[0], bug ]
                except:
                    pass
                
    for key in unique_dic:
        sheetList.append(unique_dic[key])
        
    OutPut('Bug_ExpList', sheetList)

#输出poc漏洞列表
def poc_list_export(datalist):
    unique_dic = {}
    sheetList = [['url', 'bug_poc']]

    for i in datalist:
        p = re.findall(r"\[\+]\shttp[^\s].*", i)
        # print(p)

        if len(p) != 0:
            p1 = list(p)
            for u in p1:
                # url = re.findall(r"http[^\s].*\s", u)
                url = re.findall(r"(?P<url>https?://\S+)", u)
                bug = u.replace(url[0], '').replace("[+]", "").replace('\t', '').strip()
                # url.append(bug)
                # sheetList.append(url)
                try:
                    unique_dic[f'{url[0]}_{bug}'] = [ url[0], bug ]
                except:
                    pass
                
    for key in unique_dic:
        sheetList.append(unique_dic[key])
    OutPut('Bug_PocList', sheetList)

#输出title
def http_scan_export(datalist):
    unique_dic = {}
    sheetList = [['url', 'lengh', 'code', 'title']]

    for i in datalist:
        p = re.findall(r'\[\*]\sWebTitle.*', i)

        if len(p) != 0:
            p1 = list(p)
            for u in p1:
                all_url = re.findall(r"http[^\s]+", u)
                url = [all_url[0]]
                code = re.findall(r'(?<=code:)[^\s]+', u)
                len1 = re.findall(r'(?<=len:)[^\s]+', u)
                title = re.findall(r'(?<=title:).*', u)
                
                try:
                    urlp = urlparse(url[0])
                    host = urlp.netloc.split(":")[0]
                    # unique_dic[f'{url[0]}_{code[0]}_{title}'] = [ url[0], str(len1).strip("['").strip("']'"), code[0], str(title).strip("['").strip("']'") ]
                    if unique_dic.get(host):
                        unique_dic[host][f'{url[0]}_{code[0]}_{title}'] = [ url[0], str(len1).strip("['").strip("']'"), code[0], str(title).strip("['").strip("']'") ]
                    else:
                        unique_dic[host] = {
                            f'{url[0]}_{code[0]}_{title}' : [ url[0], str(len1).strip("['").strip("']'"), code[0], str(title).strip("['").strip("']'") ]
                        }
                except:
                    traceback.print_exc()
                
    for host in unique_dic:
        for url in unique_dic[host]:
            sheetList.append(unique_dic[host][url])
        
    OutPut('Title', sheetList)

#输出弱口令
def weekpass_scan_export(datalist):
    unique_dic = {}
    sheetList = [['ip', 'port', 'protocol', 'user&passwd']]
    for i in datalist:
        try:
            p = re.findall(r'((ftp|mysql|mssql|SMB|RDP|Postgres|SSH|oracle|SMB2-shares)(:|\s).*)', i, re.I)
            rd = re.findall(r'((redis|Mongodb)(:|\s).*)', i, re.I)
            mc = re.findall(r"((Memcached)(:|\s).*)", i, re.I)

            if len(p) != 0 :
                if '[-]' in i or '[+] Product' in i:
                    continue 
                p1 = list(p)
                all = p1[0][0]
                ip = ""
                port = ""
                userAndPwd = ""
                
                all = p1[0][0].split(":")
                try:
                    userAndPwd = all[3]
                except:
                    userAndPwd = []
                    pass
                protocol = all[0]
                port = all[2]
                    
                ip = re.findall(r"\d+\.\d+\.\d+\.\d+", str(all))
                ip.append(port)
                ip.append(protocol)
                ip.append(userAndPwd)
                sheetList.append(ip)


            if len(rd) != 0 and len(rd[0][0].split(" ")) == 2:
                rd1 = list(rd)

                rd_all = rd1[0][0].split(" ")
                passwd = rd_all[-1]
                server = rd1[0][1]
                port = (rd_all[0].split(":"))[2]
                ip = re.findall(r"\d+\.\d+\.\d+\.\d+", rd1[0][0])
                ip.append(port)
                ip.append(server)
                ip.append(passwd)
                sheetList.append(ip)

            if len(mc) != 0:
                mc1 = list(mc)

                mc_all = mc1[0][0].split(" ")
                passwd = mc_all[2]
                server = mc_all[0]
                port = (mc_all[1].split(":"))[-1]
                ip = re.findall(r"\d+\.\d+\.\d+\.\d+", mc1[0][0])
                ip.append(port)
                ip.append(server)
                ip.append(passwd)
                sheetList.append(ip)
        except:
            traceback.print_exc()
            
    OutPut('WeakPasswd', sheetList)


#输出指纹信息
def fingerprint_scan_export(datalist):
    # w1 = wb.create_sheet('')
    unique_dic = {}
    sheetList = [['url', 'finger']]

    for i in datalist:
        p = re.findall(r'.*InfoScan.*', i)
        # print(p)

        if len(p) != 0:
            p1 = list(p)
            for u in p1:
                url = re.findall(r'http[^\s]+', u)
                finger = u.split(url[0])[-1].strip()
                try:
                    unique_dic[f'{url[0]}_{finger}'] = [ url[0], finger ]
                except:
                    pass
    for key in unique_dic:
        sheetList.append(unique_dic[key])
    OutPut('Finger', sheetList)

#表格输出整理
def OutPut(sheetname, sheetList):
    ws = wb.create_sheet(sheetname)
    
    # 设置前4列的默认列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 30
    if sheetname == 'OpenPort':
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
    if sheetname == 'OsList':
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 130
    elif sheetname == 'Title':
        ws.column_dimensions['D'].width = 100
    elif sheetname == 'vulnscan':
        ws.column_dimensions['A'].width = 70
        ws.column_dimensions['B'].width = 100
    elif sheetname == 'Product':
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['c'].width = 50
        ws.column_dimensions['d'].width = 90
    else:
        ws.column_dimensions['D'].width = 30
    
    
    #将列表写入sheet
    for i in sheetList:
        # 解决\x03此类特殊字符报错
        try:
            print("[+] write:", i)
            ws.append(i)
        except openpyxl.utils.exceptions.IllegalCharacterError:
            i[-1] = ILLEGAL_CHARACTERS_RE.sub(r'', i[-1])
            ws.append(i)
            traceback.print_exc()
        except Exception as e:
            print('[debug] write excel err, content:', i)
            traceback.print_exc()

    #首行格式
    for row in ws[f"A1:{chr(65 + len(output_lines[0]) - 1)}1"]:
        for cell in row:
            cell.font = Font(size=12, bold=True)
            
    
    # 设置交替的颜色
    color_header = "249978"
    color1 = "ffffff"   
    color2 = "f3f3fA"
    current_color = None
    # 设置表头的字体为微软雅黑，颜色为白色
    header_font = Font(name="微软雅黑", color="FFFFFF")
    line_count = ws.max_row
    for cell in ws[1]: # ws[1] 就是表头
        cell.font = header_font
        cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
        
    # 遍历每一行并设置颜色
    for index, row in enumerate(range(2, line_count+1)):
        if index == 0:
            current_color = color1
        else:
            if ws[row][0].value != ws[row-1][0].value:
                if current_color == color1:
                    current_color = color2
                else:
                    current_color = color1
        
        fill = PatternFill(start_color=current_color, end_color=current_color, fill_type="solid")

        for cell in ws[row]:
            cell.fill = fill
    
#输出识别出的产品信息
def product_scan_export(datalist):
    unique_dic = {}
    sheetList = [['URL', '状态码', '标题', '产品']]

    for i in datalist:
        find_products = re.findall(r'\[\+\] Product (.*)', i)

        if len(find_products) != 0:
            product_list = list(find_products)
            for product_line in product_list:
                try:
                    slice = product_line.split()
                    url = slice[0]
                    httpcode = slice[1]
                    slice_tmp = product_line.split('\t')
                    title = slice_tmp[2]
                    title = title[1:]
                    title = title[:-1]
                    product = ''
                    if len(slice_tmp) > 3:
                        product = slice_tmp[3]
                        
                    urlp = urlparse(url)
                    host = urlp.netloc.split(":")[0]
                    if unique_dic.get(host):
                        unique_dic[host][f'{url}_{httpcode}_{title}'] = [ url, httpcode, title,  product ]
                    else:
                        unique_dic[host] = {
                            f'{url}_{httpcode}_{title}' : [ url, httpcode, title,  product ]
                        }
                except:
                    traceback.print_exc()
                
    for host in unique_dic:
        for key in unique_dic[host]:
            sheetList.append(unique_dic[host][key])
        
    OutPut('Product', sheetList)



def getInput():
    if len(sys.argv) != 2:
        print("\n[*] fscan结果整理脚本，输出为.xlsx文件\n\nUsage: \n\n    python3 FscanOutput_v1.02.py result.txt\n")
        exit()
    if not os.path.exists(sys.argv[1]):
        print(f"[{sys.argv[1]}] 文件不存在")
        exit()
    return sys.argv[1]

if __name__ == "__main__":

    print(r'''

============================================================
 ______                    ____        _               _   
|  ____|                  / __ \      | |             | |  
| |__ ___  ___ __ _ _ __ | |  | |_   _| |_ _ __  _   _| |_ 
|  __/ __|/ __/ _` | '_ \| |  | | | | | __| '_ \| | | | __|
| |  \__ \ (_| (_| | | | | |__| | |_| | |_| |_) | |_| | |_ 
|_|  |___/\___\__,_|_| |_|\____/ \__,_|\__| .__/ \__,_|\__|
                                          | |              
                                          |_|           Plus       
============================================================
                                         ---By zoro123 v2.2
           ''')
    output_lines, output_data = OpenFile()

    wb = openpyxl.Workbook()
    open_port_export(output_lines)
    alive_ip_export(output_lines)
    exp_list_export(output_lines)
    poc_list_export(output_lines)
    os_scan_export(output_lines)
    http_scan_export(output_lines)
    weekpass_scan_export(output_lines)
    fingerprint_scan_export(output_lines)
    vulnscan_export(output_lines)
    product_scan_export(output_lines)
    
    ws5 = wb["Sheet"]
    wb.remove(ws5)
    input_filename = sys.argv[1].split(".txt")[0]
    Output_xlsx = (f"%s_{time.strftime('%Y-%m-%d_%H-%M-%S', time.localtime())}.xlsx" % input_filename)
    # wb.save(f"%s_{time.strftime('%Y-%m-%d_%H-%M-%S', time.localtime())}.xlsx" %input_filename)
    wb.save(Output_xlsx)
    print("[+]文件读取成功，处理结果如下······\n")
    New_fscanxlsx = p.load_workbook(Output_xlsx)

    print("+---------------------------------+\n")
    wt = New_fscanxlsx['OpenPort']
    print("[+++]探测存活端口共计：%s 个" % (wt.max_row-1))
    wt = New_fscanxlsx['AliveIp']
    print("[+++]探测存活IP段共计：%s 个" % (wt.max_row - 1))
    wt1 = New_fscanxlsx['Bug_ExpList']
    print("[+++]Exp可利用漏洞共计：%s 个" % (wt1.max_row-1))
    wt2 = New_fscanxlsx['Bug_PocList']
    print("[+++]Poc可利用漏洞共计：%s 个" % (wt2.max_row-1))
    wt3 = New_fscanxlsx['OsList']
    print("[+++]成功识别操作系统共计：%s 个" % (wt3.max_row-1))
    wt4 = New_fscanxlsx['Title']
    print("[+++]成功探测Web服务共计：%s 条" % (wt4.max_row-1))
    wt5 = New_fscanxlsx['WeakPasswd']
    print("[+++]成功破解账号密码共计：%s 个" % (wt5.max_row-1))
    wt6 = New_fscanxlsx['Finger']
    print("[+++]成功识别指纹共计：%s 个" % (wt6.max_row-1))
    wt7 = New_fscanxlsx['vulnscan']
    print("[+++]存在漏洞的资产共：%s 个" % (wt7.max_row-1))
    print("+---------------------------------+\n")

    print('[+]结果已经整理输出至 -- %s -- 文件所在目录！\n' % sys.argv[1])
    print('--> 文件名为：%s' % Output_xlsx)










